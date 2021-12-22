import os
import csv
import json
from pathlib import Path
from openpyxl import load_workbook

from preprocess_data import *


def precheck(poll=None) -> bool:
    """
      Function to precheck if poll data is available or not.
    """
    messages = create_poll_messages(poll)
    paths = create_poll_paths(poll)

    if poll is None:
        print(messages["no_poll_id"])
        return False
    else:
        # Check if poll subfolder exists
        if not os.path.isdir(paths["poll"]):
            print(messages["no_poll_folder"])
            return False
        else:
            # Create file paths
            codebook = Path(paths["codebook"])
            data = Path(paths["data"])
            # Check if poll files exist
            if not codebook.exists():
                print(messages["no_codebook"])
                return False
            elif not data.exists():
                print(messages["no_data"])
                return False
            else:
                # Everyting there
                return True


def save_results(poll, name, data) -> None:

    paths = create_target_paths(poll, name)

    if not os.path.isdir(paths["target_dir"]):
        # Create target folder
        os.mkdir(paths["target_dir"])

    with open(paths["target_file"], "w+", encoding="utf-8") as file:
        json.dump(
            data,
            file,
            ensure_ascii=False,
            indent=4
        )


def clean_cell(cell) -> str:
    string = str(cell)
    str_list = string.split("  ")
    str_list = filter(lambda i: i != "", str_list)
    string = " ".join(str_list)
    string = string.strip()
    string = string.replace("  ", " ")
    return string


def preprocess_data(poll, delimiter) -> None:

    paths = create_poll_paths(poll)
    data_path = paths["data"]

    data = []

    with open(data_path, encoding="utf-8") as csv_file:
        csv_reader = csv.DictReader(csv_file, delimiter=delimiter)
        for row in csv_reader:
            data.append(row)

    reduced_data = []

    # Reduce data
    for obj in data:
        new_obj = reduce_data_obj(poll, obj)
        reduced_data.append(new_obj)

    # Downsize to the first, random, 1000 people
    reduced_data = sorted(reduced_data, key=lambda obj: obj["SECURITYID"])
    reduced_data = reduced_data[0:1000]

    for obj in reduced_data:
        # Take care of special character in "CODERESP"
        if "﻿CODERESP" in obj:
            coderesp = obj["﻿CODERESP"]
            del obj["﻿CODERESP"]
            obj["CODERESP"] = coderesp

        # Take care of wrong CITIZENREPA codes
        if obj["CITIZENREPA"] == "9":
            obj["CITIZENREPA"] = "8"

        if obj["CONFESSR"] == " ":
            obj["CONFESSR"] = "99999998"

    save_results(poll, "data", reduced_data)


def reduce_data_obj(poll, obj):
    """
    Removes the data for the other votes from the same day as the current poll.
    """

    poll_info = poll_vote_info()
    alphabet = abc()

    vote_nr = poll_info[poll]["vote"]
    nr_of_votes = poll_info[poll]["nr_of_votes"]
    letter = alphabet[vote_nr - 1]

    votes_to_remove = list(range(1, nr_of_votes + 1))
    votes_to_remove = list(filter(lambda nr: nr != vote_nr, votes_to_remove))
    letters_to_remove = list(filter(lambda l: l != letter, alphabet))

    keys_to_remove_complete = []
    for v in votes_to_remove:
        keys_to_remove_complete.append(keys_to_remove_by_nr(v))

    for l in letters_to_remove:
        keys_to_remove_complete.append(keys_to_remove_by_letter(l))
    # Flatten keys_to_remove
    keys_to_remove_complete = [
        key for key_list in keys_to_remove_complete for key in key_list]

    keys_to_remove_start = [
        keys_to_remove_by_start(v) for v in votes_to_remove]
    # Flatten starts_with_to_remove
    keys_to_remove_start = [
        key for key_list in keys_to_remove_start for key in key_list]

    keys_to_remove_complete = [key.upper() for key in keys_to_remove_complete]
    keys_to_remove_start = [key.upper() for key in keys_to_remove_start]

    # Remove unwanted keys directly
    for key in keys_to_remove_complete:
        obj.pop(key, False)

    # Remove unwanted keys complicated
    keys_to_delete_by_start = []
    for key in obj:
        for k in keys_to_remove_start:
            if key.startswith(k):
                keys_to_delete_by_start.append(key)

    for key in keys_to_delete_by_start:
        obj.pop(key, False)

    return obj


def preprocess_codebook(poll) -> None:
    """
      Function to preprocess post-vote poll data. 
      The functions reads the codebook of a specific post-vote poll
      and generates corresponding arrangements.json and selections.json files. 
      The function requests a swissvote vote number parameter to find
      post-vote poll data in a corresponding subfolder.
      :param poll: int
    """

    paths = create_poll_paths(poll)

    wb = load_workbook(filename=paths["codebook"])
    codebook = wb[paths["codebook_sheet"]]

    columns = column_info()

    min_row = codebook.min_row
    max_row = codebook.max_row

    selections = {}
    arrangements = {}

    curr_super_sel = "_-_"
    curr_sel = ""
    curr_att = ""
    is_super_selection = False
    is_selection = False

    def create_selection(cell) -> None:
        if curr_sel.startswith(curr_super_sel):
            selections[curr_super_sel]["selections"][curr_sel][columns["lang_columns"]
                                                               [cell.column]] = clean_cell(cell.value)
        else:
            selections[curr_sel][columns["lang_columns"]
                                 [cell.column]] = clean_cell(cell.value)

    def create_arrangement(cell) -> None:
        arrangements[curr_sel][curr_att][columns["lang_columns"]
                                         [cell.column]] = clean_cell(cell.value)

    # Iterate rows in codebook (starting from 2nd row)
    for row in codebook.iter_rows(
        min_row=min_row + 1,
        max_row=max_row,
        min_col=columns["min_col"],
        max_col=columns["max_col"]
    ):

        if row[1].value in skip_indicators():
            # Skip row
            continue

        if str(row[2].value).strip() in skip_indicators():
            if curr_sel not in continous_indicators():
                # Try to remove already added keys from dicts
                selections.pop(curr_sel, False)
                arrangements.pop(curr_sel, False)
            # Skip row
            continue

        # Iterate cells in codebook row
        for cell in row:

            if cell.column == 1 and cell.value != None:
                # It's a super selection
                curr_super_sel = cell.value.upper()
                # Already add super selection to dict
                selections[curr_super_sel] = {
                    "de": row[columns["de_col"] - 1].value,
                    "fr": row[columns["fr_col"] - 1].value,
                    "it": row[columns["it_col"] - 1].value,
                    "selections": {}
                }
            elif cell.column == 1 and cell.value == None:
                # Skip cell
                continue

            elif cell.column == 2:
                # It's a code!

                cell_value = str(cell.value)

                # Get propper keys
                cell_value = cell_value.split(" ")[0]
                cell_value = cell_value.split("-")[0]

                if cell_value.isnumeric() == False and cell.value != None:
                    # It's a selection
                    is_super_selection = False
                    is_selection = True
                    # Set current selection
                    curr_sel = cell_value.upper()

                    # Create keys in final objects
                    if curr_sel.startswith(curr_super_sel):
                        selections[curr_super_sel]["selections"][curr_sel] = {}
                    else:
                        selections[curr_sel] = {}

                    if curr_sel not in continous_indicators():
                        arrangements[curr_sel] = {}

                elif cell_value.isnumeric() == True and cell_value != None:
                    # It's a selection attribute
                    is_super_selection = False
                    is_selection = False
                    # Set current attribute
                    curr_att = cell_value
                    # Create subkey in final arrangements object
                    arrangements[curr_sel][cell_value] = {}

                else:
                    # It's a super selection!
                    # See if cell.column == 1 and cell.value != None
                    is_super_selection = True
                    # Skip cell
                    continue

            else:
                # It's a label!
                if not is_super_selection:
                    if is_selection:
                        # It's a selection key
                        create_selection(cell)
                    else:
                        # It's a selection attribute
                        create_arrangement(cell)

    # Remove controls and weightings
    for item in delete_indicators():
        selections.pop(item, False)
        arrangements.pop(item, False)

    # Reduce data!
    selections = reduce_data_obj(poll, selections)
    arrangements = reduce_data_obj(poll, arrangements)

    # Take care of wrong lrsp values
    arrangements["LRSP"] = {
        "0": {
            "de": "0 ganz links",
            "fr": "0 tout à gauche",
            "it": "0 estrema sinistra"
        },
        "1": {
            "de": "1",
            "fr": "1",
            "it": "1"
        },
        "2": {
            "de": "2",
            "fr": "2",
            "it": "2"
        },
        "3": {
            "de": "3",
            "fr": "3",
            "it": "3"
        },
        "4": {
            "de": "4",
            "fr": "4",
            "it": "4"
        },
        "5": {
            "de": "5 Zentrum",
            "fr": "5 centre",
            "it": "5 centro"
        },
        "6": {
            "de": "6",
            "fr": "6",
            "it": "6"
        },
        "7": {
            "de": "7",
            "fr": "7",
            "it": "7"
        },
        "8": {
            "de": "8",
            "fr": "8",
            "it": "8"
        },
        "9": {
            "de": "9",
            "fr": "9",
            "it": "9"
        },
        "10": {
            "de": "10 ganz rechts",
            "fr": "10 tout à droite",
            "it": "10 estrema destra"
        },
        "98": {
            "de": "weiss nicht / keine Angabe",
            "fr": "ne sais pas / aucune indication",
            "it": "non so / nessuna risposta"
        }
    }

    # Save results
    save_results(poll, "selections", selections)
    save_results(poll, "arrangements", arrangements)


if __name__ == "__main__":
    for poll in poll_ids():
        print("")
        print(f"Checking files for poll {poll}")
        if precheck(poll):
            print(f"Preprocessing codebook of poll {poll}")
            preprocess_codebook(poll)
            print(f"Preprocessing data of poll {poll}")
            delimiter = ","
            if poll in polls_with_semicolon():
                delimiter = ";"
            preprocess_data(poll, delimiter)
